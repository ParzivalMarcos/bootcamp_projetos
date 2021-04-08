import os
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By


class WebScraping:
    def __init__(self):
        chrome_options = Options()
        chrome_options.binary_location = os.getcwd() + os.sep + 'chrome-win' + \
            os.sep + 'chrome.exe'
        self.driver = webdriver.Chrome(executable_path=r'chromedriver.exe', \
            options=chrome_options)
        

    def Iniciar(self):
        self.proxima_pagina = 1
        self.cria_cabecalho_planilha()
        self.acessar_site()

    def cria_cabecalho_planilha(self):
        self.planilha = openpyxl.Workbook()
        self.planilha.create_sheet('Celulares')
        self.planilha_celuares = self.planilha['Celulares']
        self.planilha_celuares.cell(row=1, column=1, value='Marca')
        self.planilha_celuares.cell(row=1, column=2, value='Preco')

        # self.planilha.save('Planilha_teste.xlsx')

    def acessar_site(self):
        self.driver.get('https://telefonesimportados.netlify.app')
        self.captura_informacoes()


    def captura_informacoes(self):
        # Tag xpath para produtos: //h2/a
        # Tag xpath para preos: //div/ins
        produtos = self.driver.find_elements(By.XPATH, '//h2/a')
        precos = self.driver.find_elements(By.XPATH, '//div/ins')
        
        self.gravar_informacoes(produtos, precos)  
        self.navegar_proximo_link()      


    def gravar_informacoes(self, produtos, precos):
        for item in range(0, len(produtos)):
            dados = [
                produtos[item].text,
                precos[item].text,
            ]
            self.planilha_celuares.append(dados)

    def navegar_proximo_link(self):
        self.proxima_pagina += 1
        if self.proxima_pagina > 5:
            self.driver.quit()
            self.planilha.save('produtos_e-commerce.xlsx')
        else:
            link = f'https://telefonesimportados.netlify.app/shop{self.proxima_pagina}.html'
            self.driver.get(link)
            self.captura_informacoes()   
