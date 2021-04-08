import os
import openpyxl


def criar_paginas():
    nome_pagina = input('Por favor, insira o nome da pagina\n-> ')
    planilha.create_sheet(str(nome_pagina))


def criar_colunas():
    coluna_planilha = []
    pagina = input('Qual página deseja editar?\n-> ')

    pagina_planilha = planilha[pagina]
    coluna_nome = input('Digite o nome da coluna que deseja criar\n-> ')
    coluna_planilha.append(coluna_nome)

    pagina_planilha.append(coluna_planilha)


def salvar_planilha(planilha):
    nome_planilha = input('Digite o nome do arquivo a ser salvo: ')
    planilha.save(f'{nome_planilha}.xlsx')
    print(f'Planilha {nome_planilha}.xlsx salva com sucesso !!')


if __name__ == '__main__':
    planilha = openpyxl.Workbook()

    print('Bem vindo ao gerador de planilhas !!')
    print('Vamos criar uma nova pagina !!')

    # continua = True
    criar_paginas()
    while True:        

        escolha = input(str('Deseja criar mais uma pagina? [sim/nao]\n-> ')).lower()

        if escolha == 'sim':
            criar_paginas()
        else:
            print(f'As paginas criadas são\n -> {planilha.sheetnames}')
        
            break

    criar_colunas()    
    while True:

        escolha = input('Deseja criar mais uma coluna? [sim/nao]\n-> ')
        if escolha == 'sim':
            criar_colunas()
        else:
            print('As colunas foram criadas !!')
        
            break


    salvar_planilha(planilha)


    # print('Digite o nome da coluna')

    # print('Deseja criar mais colunas?')

    # if True:
    #     print('Digite o nome da coluna')
    # else:
    #     print('Digite os valores a incluir nas colunas separados por virgulas')

    # print('Digite o nome do arquivo a ser salvo')

    '''
    Capurar o nome da pagina;
        Perguntar se deseja criar mais uma pagina;
    Mostrar as paginas criadas;
    Perguntar qual pagina para editar;
    Capturar o nome das colunas;
        Perguntar se deseja continuar a criar colunas;
    Capturar os dados para incluir nas linhas das colunas.

    Finalizar o programa
    '''